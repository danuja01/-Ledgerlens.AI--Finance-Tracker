#!/usr/bin/env python3
"""Export workbooks from data/ to src/data/*.json (see data/ next to this package)."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT_JSON = ROOT / "src" / "data"
# Source workbooks live next to the app (commit or gitignore as you prefer)
SOURCE = ROOT / "data"

MAPPING_XLSX = SOURCE / "final_subcategory_mapping.xlsx"
# Prefer a stable name; fall back to a dated export if present
_TX_CANDIDATES = (
    SOURCE / "transactions.xlsx",
    SOURCE / "2026-04-19.xlsx",
)
TX_XLSX = next((p for p in _TX_CANDIDATES if p.exists()), _TX_CANDIDATES[0])

# Display names from build prompt + fallbacks from mapping keys
DISPLAY_NAME_OVERRIDES: dict[str, str] = {
    "🍜 Food": "Food & Groceries",
    "💾 Subscriptions": "Subscriptions",
    "🚖 Transport": "Transport",
    "💄 Beauty": "Beauty & Care",
    "🏠 House Rent": "House Rent",
    "🏥 Health": "Health & Wellness",
    "💰 Salary": "Salary",
    "🧾 Payments": "Payments",
    "🛒 Tech Goods": "Tech & Devices",
    "👬🏻 Social Life": "Social & Lifestyle",
    "👩‍❤️‍👨 Parents": "Family Support",
    "🌟 Investment": "Investments",
    "💰Cash Withdraw": "Cash Withdrawals",
    "🏦 Loan": "Loans",
    "🧥 Apparel": "Apparel",
    "🪑 Household": "Household",
    "📙 Education": "Education",
    "🤑 Allowance": "Allowance",
    "💸 Freelance": "Freelance Income",
    "🎁 Gift": "Gifts",
    "🎁 Donation": "Donations",
    "📈 TAX": "Taxes",
    "🧘🏼 Health": "Health",
    "🚗 Vehicle Maintain": "Vehicle Maintenance",
    "📍 Co Work": "Co-working",
}


def split_merged_notes(text: str | None) -> list[str]:
    if not text or not str(text).strip():
        return []
    # Excel uses commas and semicolons
    parts = re.split(r"[,;]", str(text))
    return [p.strip() for p in parts if p and str(p).strip()]


def normalize_key(s: str) -> str:
    return re.sub(r"\s+", " ", s.lower().strip())


def export_mapping() -> dict:
    wb = openpyxl.load_workbook(MAPPING_XLSX, read_only=True, data_only=True)
    ws = wb["Final Mapping"]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        main = row[0]
        if main is None:
            continue
        main_s = str(main).strip()
        sub_cell = row[1]
        merged = row[2]
        sub_display = str(sub_cell).strip() if sub_cell is not None else ""
        variants = split_merged_notes(merged)
        if not sub_display and variants:
            sub_display = variants[0]
        if main_s not in out:
            disp = DISPLAY_NAME_OVERRIDES.get(main_s, main_s)
            out[main_s] = {"displayName": disp, "subCategories": {}}
        bucket = out[main_s]["subCategories"]
        for v in variants:
            k = normalize_key(v)
            if k:
                bucket[k] = sub_display if sub_display else v
    wb.close()
    # Ensure displayName for any key only in data
    for k, v in out.items():
        if v["displayName"] == k and k in DISPLAY_NAME_OVERRIDES:
            v["displayName"] = DISPLAY_NAME_OVERRIDES[k]
    return out


def direction_from_ie(val: str | None) -> str:
    if not val:
        return "expense"
    v = str(val).strip()
    if v in ("Income", "Income Balance"):
        return "income"
    if v == "Exp.":
        return "expense"
    if v == "Transfer-In":
        return "income"
    if v == "Transfer-Out":
        return "expense"
    return "expense"


def iso_date(d: datetime | date | None) -> str:
    if d is None:
        return "1970-01-01"
    if isinstance(d, datetime):
        return d.date().isoformat()
    if isinstance(d, date):
        return d.isoformat()
    return "1970-01-01"


def period_str(p: object) -> str:
    if isinstance(p, datetime):
        return p.isoformat()
    if isinstance(p, date):
        return datetime.combine(p, time.min).isoformat()
    return str(p)


def export_transactions() -> list[dict]:
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_out: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or row[0] is None:
            continue
        period = row[0]
        category = row[2]
        note = row[4]
        amount = row[5]
        ie = row[6]
        if amount is None:
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            continue
        cat_s = str(category).strip() if category is not None else ""
        note_s = str(note).strip() if note is not None else ""
        rows_out.append(
            {
                "id": f"tx-{idx}",
                "date": iso_date(period),
                "period": period_str(period),
                "category": cat_s,
                "note": note_s,
                "amount": round(amt, 2),
                "direction": direction_from_ie(ie),
            }
        )
    wb.close()
    return rows_out


def main() -> None:
    OUT_JSON.mkdir(parents=True, exist_ok=True)
    SOURCE.mkdir(parents=True, exist_ok=True)
    if not MAPPING_XLSX.is_file():
        print(f"Missing mapping workbook: {MAPPING_XLSX}", file=sys.stderr)
        sys.exit(1)
    if not TX_XLSX.is_file():
        print(
            f"Missing transactions workbook. Place one of:\n"
            f"  {SOURCE / 'transactions.xlsx'}\n"
            f"  {SOURCE / '2026-04-19.xlsx'}",
            file=sys.stderr,
        )
        sys.exit(1)
    mapping_path = OUT_JSON / "categoryMapping.json"
    tx_path = OUT_JSON / "transactions.json"
    mapping = export_mapping()
    with mapping_path.open("w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)
    print(f"Wrote {mapping_path} ({len(mapping)} categories)")
    txs = export_transactions()
    with tx_path.open("w", encoding="utf-8") as f:
        json.dump(txs, f, ensure_ascii=False, indent=2)
    print(f"Wrote {tx_path} ({len(txs)} transactions)")


if __name__ == "__main__":
    main()
